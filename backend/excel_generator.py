from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

# Column definitions: (header_name, json_key, width)
COLUMNS = [
    ("Country", "country", 14),
    ("City", "city", 14),
    ("Retailer", "retailer", 16),
    ("Store Format", "store_format", 14),
    ("Store Name", "store_name", 18),
    ("Photo", "photo", 35),
    ("Shelf Location", "shelf_location", 22),
    ("Shelf Levels", "shelf_levels", 12),
    ("Shelf Level", "shelf_level", 12),
    ("Product Type", "product_type", 14),
    ("Branded/Private Label", "branded_private_label", 20),
    ("Brand", "brand", 18),
    ("Sub-brand", "sub_brand", 16),
    ("Product Name", "product_name", 22),
    ("Flavor", "flavor", 28),
    ("Facings", "facings", 10),
    ("Price (Local Currency)", "price_local", 20),
    ("Currency", "currency", 10),
    ("Price (EUR)", "price_eur", 12),
    ("Packaging Size (ml)", "packaging_size_ml", 18),
    ("Price per Liter (EUR)", "price_per_liter_eur", 20),
    ("Need State", "need_state", 14),
    ("Juice Extraction Method", "juice_extraction_method", 22),
    ("Processing Method", "processing_method", 18),
    ("HPP Treatment", "hpp_treatment", 14),
    ("Packaging Type", "packaging_type", 16),
    ("Claims", "claims", 35),
    ("Bonus/Promotions", "bonus_promotions", 22),
    ("Stock Status", "stock_status", 14),
    ("Est. Linear Meters", "est_linear_meters", 18),
    ("Fridge Number", "fridge_number", 14),
    ("Confidence Score", "confidence_score", 16),
    ("Notes", "notes", 35),
]

# Styles
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
DATA_FONT = Font(name="Arial", size=10)
LIGHT_GREY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
NO_FILL = PatternFill(fill_type=None)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_FONT = Font(name="Arial", size=10, bold=True, color="FF0000")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
HEADER_BORDER = Border(
    left=Side(style="thin", color="1F3864"),
    right=Side(style="thin", color="1F3864"),
    top=Side(style="thin", color="1F3864"),
    bottom=Side(style="thin", color="1F3864"),
)


def generate_excel(skus: list[dict], output_path: str) -> str:
    """Generate a formatted Excel file from SKU data. Returns the output file path."""
    wb = Workbook()
    ws = wb.active
    ws.title = "SKU Data"

    # Freeze the header row
    ws.freeze_panes = "A2"

    # Write header row
    for col_idx, (header, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = HEADER_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Header row height
    ws.row_dimensions[1].height = 30

    # Price EUR column letter and Packaging Size column letter (for formula)
    price_eur_col = None
    packaging_ml_col = None
    price_per_liter_col = None
    confidence_col = None
    stock_col = None

    for col_idx, (_, key, _) in enumerate(COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        if key == "price_eur":
            price_eur_col = col_letter
        elif key == "packaging_size_ml":
            packaging_ml_col = col_letter
        elif key == "price_per_liter_eur":
            price_per_liter_col = col_letter
        elif key == "confidence_score":
            confidence_col = col_idx
        elif key == "stock_status":
            stock_col = col_idx

    # Write data rows
    for row_idx, sku in enumerate(skus, start=2):
        is_alt_row = (row_idx % 2 == 0)
        row_fill = LIGHT_GREY_FILL if is_alt_row else NO_FILL

        for col_idx, (_, key, _) in enumerate(COLUMNS, start=1):
            col_letter = get_column_letter(col_idx)

            # Price per Liter: use Excel formula
            if key == "price_per_liter_eur":
                formula = f'=IF(AND({price_eur_col}{row_idx}<>"",{packaging_ml_col}{row_idx}<>""),{price_eur_col}{row_idx}/({packaging_ml_col}{row_idx}/1000),"")'
                cell = ws.cell(row=row_idx, column=col_idx, value=formula)
            else:
                value = sku.get(key, "")
                if value is None:
                    value = ""
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

            cell.font = DATA_FONT
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = THIN_BORDER

        # Conditional formatting: Confidence Score
        if confidence_col:
            conf_cell = ws.cell(row=row_idx, column=confidence_col)
            conf_val = sku.get("confidence_score", 0)
            if isinstance(conf_val, (int, float)):
                if conf_val >= 75:
                    conf_cell.fill = GREEN_FILL
                elif conf_val >= 55:
                    conf_cell.fill = YELLOW_FILL
                else:
                    conf_cell.fill = RED_FILL

        # Conditional formatting: Stock Status
        if stock_col:
            stock_cell = ws.cell(row=row_idx, column=stock_col)
            stock_val = sku.get("stock_status", "")
            if stock_val == "Out of Stock":
                stock_cell.fill = RED_FILL
                stock_cell.font = RED_FONT

        # Row height
        ws.row_dimensions[row_idx].height = 20

    # Format Price columns as numbers
    for col_idx, (_, key, _) in enumerate(COLUMNS, start=1):
        if key in ("price_local", "price_eur", "packaging_size_ml", "est_linear_meters"):
            for row_idx in range(2, len(skus) + 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00' if 'price' in key else '#,##0'

    # Save
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
